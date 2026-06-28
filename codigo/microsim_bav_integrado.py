#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
microsim_bav_integrado.py
==========================
Microsimulacion temporal+fiscal integrada del modelo bAV adaptado a Espana.
Reescritura completa (junio 2026) que sustituye a bav_simulacion_v3.py
(descartado: autoria/cita fabricadas, cadena de Markov sintetica, parametros
SS/GVF sin trazabilidad) y temporaliza microsim_gini_bav.py (foto fija 2023)
sobre 30 anios (2023-2053), siguiendo las 4 decisiones de JMLZ del 2026-06-25:

  1. Horizonte: 30 anios fijos 2023-2053, poblacion real por tramo/anio/sexo
     tomada de Markov_Cotizantes_Hombres_por_Tramo_2023_2053_v1.xlsx.
  2. Impacto fiscal: tipos efectivos de IRPF reales (AEAT 2023, tablas 505/587),
     no el 31% sin trazar de 04-MODULOS.md.
  3. Un unico pipeline integrado (pension con/sin bAV, fondo privado, impacto
     fiscal, Gini, sostenibilidad SS, compensacion viable) -- sin separar en
     Modulo 3 / Modulo 4 como en el paper antiguo.
  4. TAKE_UP no nulo en T1-T4 (2%/3%/5%/7%), resto sin cambios (10/25/40/45/50/50%).

HALLAZGO Y CORRECCION DE DATOS (2026-06-25):
  Markov_Cotizantes_Mujeres_por_Tramo_2023_2053_v1.xlsx tiene el Tramo 1
  construido con la matriz P_movilidad (100% de salida en un paso) en vez de
  P_obs (la matriz de movilidad real, gradual, que SI se usa para Hombres).
  Resultado: Tramo 1 Mujeres cae a 0 cotizantes en 2024 y se queda ahi todo
  el horizonte -- un artefacto, no un resultado real.

HOMOGENEIZACION DEL MECANISMO POBLACIONAL (2026-06-26, a peticion de JMLZ
tras auditoria interna -- ):
  Se detecto que la serie de Hombres ("fiable, tal cual") procedia en realidad
  de una simulacion individual (Markov_Simulacion_Hombres_2023_2053_v1.py)
  cuyo propio docstring afirma aplicar "movilidad intrinseca (P_movilidad)" y
  "descensos estructurales (f_edad)", pero en el codigo real esas dos
  funciones se definen y JAMAS se invocan en el bucle de simulacion: solo se
  aplican inflacion (reasignacion deterministica de tramo), jubilacion (edad
  >=67) y mortalidad (tabla INE). Es decir, Hombres NO usaba el mismo
  mecanismo estocastico de movilidad entre tramos (P_obs) que Mujeres, pese a
  que la Ecuacion 1 del texto describe un mecanismo unico para ambos sexos.
  Se homogeneiza aqui sustituyendo ambas cargas por un unico procedimiento,
  simetrico, aplicado por igual a Hombres y Mujeres: distribucion real de
  2023 por tramo (fuente AEAT/Markov, sin cambios) + salidas reales anuales
  (Salidas_Anuales: fallecidos + jubilados, hoja ya presente en ambos
  archivos Excel) + movilidad P_obs (propia de cada sexo, hoja "P_obs" en
  Matrices_Transicion_Hombres/Mujeres.xlsx), año a año. Se ELIMINA el
  reescalado artificial de Mujeres al ratio total de Hombres: al generarse
  ambas series con el mismo mecanismo cerrado (cohorte sin nuevas
  incorporaciones, solo salidas + movilidad interna), no hay ya un total
  "ancla" externo mas fiable al que reescalar -- cada serie es autoconsistente
  con sus propios datos de entrada.
  LIMITACION DECLARADA (antes y despues de la homogeneizacion, sin cambios):
  el modelo es de cohorte cerrada -- no incorpora la entrada de nuevos
  cotizantes jovenes al mercado de trabajo durante el horizonte 2023-2053,
  por lo que el total poblacional de ambos sexos decrece mecanicamente por
  jubilacion y mortalidad sin reposicion. Esto ya era cierto en la version
  anterior (la simulacion individual de Hombres tambien describia "una
  cohorte cerrada" en su propio docstring); la homogeneizacion no introduce
  esta limitacion, la hace simplemente explicita y simetrica para ambos sexos.

FUENTES (reales, verificables, ya en junio26/01_datos_base y 02_poblacion_dinamica):
  - Poblacion activa por tramo/sexo/anio: Markov_Cotizantes_Hombres/Mujeres_*.xlsx
  - Movilidad inter-tramo: Matrices_Transicion_Hombres/Mujeres.xlsx (hoja P_obs)
  - Renta media y tipo efectivo IRPF por tramo: Tabla505_2023.xlsx (base liquidable)
    y Tabla587_2023.xlsx (cuota liquida) -- AEAT, Estadistica IRPF 2023.
  - Cotizacion SS contingencias comunes 2023: 28,3% (23,6% empresa + 4,7%
    trabajador), tipo legal vigente (no el 31% sin origen de 04-MODULOS.md).
  - Mecanismo de Equidad Intergeneracional (MEI, Ley 21/2021): 0,6% adicional
    sobre la misma base de cotizacion por contingencias comunes (0,5% empresa
    + 0,1% trabajador), Orden PCM/74/2023, BOE-A-2023-2472. Corregido el
    2026-06-26 (dictamen revisor ronda 3): el tipo total aplicado a la base
    exenta pasa de 28,3% a 28,9%, ya que la conversion salarial al esquema
    bAV reduce la misma base de cotizacion sobre la que se calcula tambien
    el MEI, no solo la cuota de contingencias comunes.

NO SE REUTILIZA ninguna cifra de 01-OBJETIVOS.md / 04-MODULOS.md (aspiracionales,
no verificadas) ni nada de bav_simulacion_v3.py (en junio26/_excluidos/).
"""

import gc
import json
from pathlib import Path
import numpy as np
import openpyxl
import pandas as pd

# NOTA DE EMPAQUETADO PARA EL REPOSITORIO PUBLICO (no presente en la copia
# canonica del pipeline, que usa rutas absolutas del entorno de trabajo original):
# BASE y OUT se han hecho relativos a la ubicacion de este script, asumiendo la
# estructura de carpetas del repositorio (ver README.md): este archivo vive en
# codigo/, y junto a el, como hermanas, datos_entrada/ y resultados/.
# Este repositorio incluye datos_entrada/03_matrices_transicion/ (las matrices
# P_obs reales) pero NO incluye datos_entrada/01_datos_base/ ni
# datos_entrada/02_poblacion_dinamica/ (las tablas AEAT/INE y los ficheros de
# poblacion dinamica derivados de ellas): son ficheros de fuentes oficiales de
# acceso publico, no redistribuidos aqui por volumen y por no ser de elaboracion
# propia; vease el README.md para los enlaces a las fuentes originales y la
# Declaracion de disponibilidad de datos y codigo del paper. Sin esos dos
# subdirectorios, este script NO puede ejecutarse de extremo a extremo; los
# resultados ya calculados con ellos se incluyen en resultados/ para su
# verificacion directa, sin necesidad de reproducir la ejecucion.
BASE = Path(__file__).resolve().parent.parent / "datos_entrada"
OUT  = Path(__file__).resolve().parent.parent / "resultados"
OUT.mkdir(parents=True, exist_ok=True)

AÑOS = list(range(2023, 2054))   # 31 valores, 2023-2053
N_AÑOS = len(AÑOS)

# ─── Parametros de la microsimulacion individual (de microsim_gini_bav.py, ──
# ─── validados, sin cambios salvo lo indicado en las 4 decisiones) ──────────
APORT_TRAB = 0.08
TAU_EMP    = 0.15
CAP_APORT  = 4_704.0
R_BASE     = 0.020
C_GES      = 0.004
# FA_H=20 / FA_M=22 (microsim_gini_bav.py original) ERAN EXOGENOS, sin derivar
# de tablas de mortalidad reales -- corregido el 2026-06-25 (ver modulo actuarial
# mas abajo): se sustituyen por un factor de conversion a renta vitalicia
# (anualidad inmediata, cohorte, tipo tecnico 2%) calculado a partir de las
# tablas INE 36774_H/M.xlsx, especifico para el año de jubilacion de cada
# corte transversal (FACTOR_RENTA[año]['H'|'M'|'unisex']).

# Decision 4: TAKE_UP corregido (T1-T4 no nulos)
TAKE_UP = {1: 0.02, 2: 0.03, 3: 0.05, 4: 0.07,
           5: 0.10, 6: 0.25, 7: 0.40, 8: 0.45, 9: 0.50, 10: 0.50}

DELTA_T_H = {1: 0.866, 2: 0.866, 3: 0.866, 4: 0.949, 5: 0.949, 6: 0.949,
             7: 1.000, 8: 1.000, 9: 1.000, 10: 1.000}
DELTA_T_M = {1: 0.837, 2: 0.837, 3: 0.837, 4: 0.938, 5: 0.938, 6: 0.938,
             7: 1.000, 8: 1.000, 9: 1.000, 10: 1.000}

# Decision 2: tipo de cotizacion SS contingencias comunes 2023 (legal, verificable)
TASA_SS_CC  = 0.283     # 23,6% empresa + 4,7% trabajador (TGSS, Orden PCM 2023)
# Correccion 2026-06-26 (dictamen revisor ronda 3): se anade el Mecanismo de
# Equidad Intergeneracional (MEI, Ley 21/2021), 0,6% adicional sobre la misma
# base de cotizacion por contingencias comunes (Orden PCM/74/2023). La
# exencion de la base de cotizacion que genera la conversion salarial al
# esquema bAV reduce tambien la base sobre la que se calcula el MEI, por lo
# que su perdida de recaudacion debe incluirse en TASA_SS_TOTAL.
TASA_MEI    = 0.006     # 0,5% empresa + 0,1% trabajador (MEI, Orden PCM/74/2023)
TASA_SS_TOTAL = TASA_SS_CC + TASA_MEI   # 28,9% (antes 28,3%, sin MEI)

# Limites de tramo = brackets reales AEAT (Tabla505/587), no los aproximados
# de microsim_gini_bav.py (12.450/20.200/35.200/60.000)
TRAMOS_LIM = {
    1:  (-np.inf, 0),
    2:  (0, 1_500),
    3:  (1_500, 6_000),
    4:  (6_000, 12_000),
    5:  (12_000, 21_000),
    6:  (21_000, 30_000),
    7:  (30_000, 60_000),
    8:  (60_000, 150_000),
    9:  (150_000, 601_000),
    10: (601_000, np.inf),
}
TRAMO10_CAP_PRACTICO = 3_000_000.0

# ─── Carga: Tabla505 (base liquidable) + Tabla587 (cuota liquida) -> ────────
# ─── renta media y tipo efectivo IRPF reales por tramo (AEAT 2023) ──────────
def cargar_irpf_aeat():
    wb505 = openpyxl.load_workbook(BASE / "01_datos_base/Tabla505_2023.xlsx", data_only=True)
    wb587 = openpyxl.load_workbook(BASE / "01_datos_base/Tabla587_2023.xlsx", data_only=True)
    rows505 = list(wb505["Hoja1"].iter_rows(values_only=True))[10:20]   # 10 tramos (sin fila Total)
    rows587 = list(wb587["Hoja1"].iter_rows(values_only=True))[10:20]
    renta_media, tipo_efectivo = {}, {}
    for i, t in enumerate(range(1, 11)):
        base  = rows505[i][5]
        cuota = rows587[i][5]
        media = rows505[i][7]
        renta_media[t]    = max(media, 0.0)   # T1 (negativo/cero) -> renta 0 en la simulacion
        tipo_efectivo[t]  = max(cuota / base, 0.0) if base else 0.0
    return renta_media, tipo_efectivo

RENTA_MEDIA, TIPO_EFECTIVO_IRPF = cargar_irpf_aeat()


# ─── Mecanismo poblacional UNICO y homogeneizado (2026-06-26): distribucion ──
# ─── real 2023 + salidas reales anuales + movilidad P_obs propia de cada ────
# ─── sexo, aplicado de forma identica a Hombres y Mujeres -- sin reescalado ──
# ─── cruzado entre sexos. Ver HALLAZGO Y CORRECCION DE DATOS en el docstring.─
# (Adelantado respecto a versiones anteriores del script: TOTAL_H/TOTAL_M se
# necesitan ya en el modulo actuarial para ponderar el factor unisex por
# poblacion -- ver correccion 2026-06-26 tras dictamen de revisor, punto 2.3)
def _cargar_distribucion_2023_bloques(sh):
    """Lee la hoja 'Cotizantes_por_Tramo' en formato bloque ('AÑO: XXXX' +
    filas (Tramo, Cotizantes) + fila TOTAL) -- formato usado en el Excel de
    Hombres. Devuelve solo la distribucion real de 2023 por tramo (1..10)."""
    year, v0 = None, {}
    for r in sh.iter_rows(values_only=True):
        if r[0] and isinstance(r[0], str) and r[0].startswith("AÑO"):
            year = int(r[0].split(":")[1].strip())
        elif year == 2023 and isinstance(r[0], int):
            v0[r[0]] = float(r[1])
        elif year is not None and year > 2023:
            break
    return np.array([v0[t] for t in range(1, 11)])

def _cargar_salidas_bloques(sh):
    """Lee 'Salidas_Anuales' en formato bloque (mismo formato que Hombres):
    devuelve {año: {tramo: total_salidas}}, usando la columna 'Total Salidas'
    (4a columna, indice 3)."""
    year, sal = None, {}
    for r in sh.iter_rows(values_only=True):
        if r[0] and isinstance(r[0], str) and r[0].startswith("AÑO"):
            year = int(r[0].split(":")[1].strip())
            sal[year] = {}
        elif isinstance(r[0], int):
            sal[year][r[0]] = float(r[3] or 0.0)
    return sal

def _cargar_distribucion_2023_flat(sh):
    """Lee 'Cotizantes_por_Tramo' en formato plano (Año, Tramo, Descripcion,
    Cotizantes) -- formato usado en el Excel de Mujeres."""
    v0 = {}
    for r in sh.iter_rows(min_row=2, values_only=True):
        if r[0] == 2023:
            v0[r[1]] = float(r[3])
    return np.array([v0[t] for t in range(1, 11)])

def _cargar_salidas_flat(sh):
    """Lee 'Salidas_Anuales' en formato plano (Año, Tramo, Descripcion,
    Fallecidas/os, Jubiladas/os, Total_Salidas)."""
    sal = {}
    for r in sh.iter_rows(min_row=2, values_only=True):
        sal.setdefault(r[0], {})[r[1]] = float(r[5] or 0.0)
    return sal

def _cargar_P_obs(path_excel):
    wbt = openpyxl.load_workbook(path_excel, data_only=True)
    rowsp = list(wbt["P_obs"].iter_rows(min_row=2, max_row=11, max_col=11, values_only=True))
    return np.array([[r[j] for j in range(1, 11)] for r in rowsp])

def propagar_poblacion_markov(v0, sal, P_obs):
    """Mecanismo unico (homogeneizado 2026-06-26): cohorte cerrada, salidas
    reales (fallecidos+jubilados) restadas, movilidad P_obs aplicada a los
    supervivientes, año a año, SIN reescalado a ningun total externo.
    Identico procedimiento para Hombres y Mujeres -- solo cambian los datos
    de entrada (v0, sal, P_obs) propios de cada sexo."""
    pop = {2023: {t: v0[t - 1] for t in range(1, 11)}}
    v = v0.copy()
    for y in AÑOS[1:]:
        exits = np.array([sal.get(y, {}).get(t, 0.0) for t in range(1, 11)])
        survivors = np.clip(v - exits, 0, None)
        v = survivors @ P_obs
        pop[y] = {t: v[t - 1] for t in range(1, 11)}
    return pop


print("Cargando datos de entrada Hombres (distribucion 2023 + salidas + P_obs)...")
wb_h = openpyxl.load_workbook(BASE / "02_poblacion_dinamica/Markov_Cotizantes_Hombres_por_Tramo_2023_2053_v1.xlsx", data_only=True)
v0_h = _cargar_distribucion_2023_bloques(wb_h["Cotizantes_por_Tramo"])
sal_h = _cargar_salidas_bloques(wb_h["Salidas_Anuales"])
P_obs_H = _cargar_P_obs(BASE / "03_matrices_transicion/Matrices_Transicion_Hombres.xlsx")

print("Cargando datos de entrada Mujeres (distribucion 2023 + salidas + P_obs)...")
wb_m = openpyxl.load_workbook(BASE / "02_poblacion_dinamica/Markov_Cotizantes_Mujeres_por_Tramo_2023_2053_v1.xlsx", data_only=True)
v0_m = _cargar_distribucion_2023_flat(wb_m["Cotizantes_por_Tramo"])
sal_m = _cargar_salidas_flat(wb_m["Salidas_Anuales"])
P_obs_M = _cargar_P_obs(BASE / "03_matrices_transicion/Matrices_Transicion_Mujeres.xlsx")

print("Propagando poblacion Hombres y Mujeres (mecanismo homogeneizado, identico para ambos sexos)...")
POP_H = propagar_poblacion_markov(v0_h, sal_h, P_obs_H)
POP_M = propagar_poblacion_markov(v0_m, sal_m, P_obs_M)

TOTAL_H = {y: sum(POP_H[y].values()) for y in POP_H}
TOTAL_M = {y: sum(POP_M[y].values()) for y in POP_M}
print(f"  Total H 2023: {TOTAL_H[2023]:,.0f}  ->  2053: {TOTAL_H[2053]:,.0f}")
print(f"  Total M 2023: {TOTAL_M[2023]:,.0f}  ->  2053: {TOTAL_M[2053]:,.0f}")


# ─── Modulo actuarial: factor de conversion a renta vitalicia real ─────────
# (sustituye a FA_H=20/FA_M=22 exogenos -- correccion 2026-06-25 tras dictamen
# de revisor. Fuente: tablas de mortalidad proyectada INE 36774_H/M.xlsx,
# qx por edad y año 2023-2073, en por mil. Anualidad inmediata pospagable,
# nivel constante (sin indexacion, sin reversion, sin gastos de gestion --
# estos ya se cobran via C_GES en la fase de acumulacion), tipo tecnico=R_BASE
# (2%, mismo tipo que la fase de acumulacion, supuesto declarado), calculo
# por cohorte: para alguien que se jubila a los EDAD_JUB en el año A, se usa
# qx(67,A), qx(68,A+1), qx(69,A+2)... -- la diagonal de mejora proyectada de
# la propia tabla, no una tabla estatica congelada en un año.
EDAD_JUB, EDAD_MAX_TABLA = 67, 100

def _cargar_qx(nombre_archivo, hoja):
    wb = openpyxl.load_workbook(BASE / "01_datos_base" / nombre_archivo, data_only=True)
    sh = wb[hoja]
    filas = list(sh.iter_rows(values_only=True))
    años_cab = filas[0][1:]
    qx = {}
    for r in filas[1:]:
        edad = 100 if "100" in r[0] else int(r[0].split()[0])
        qx[edad] = {año: (val / 1000.0) for año, val in zip(años_cab, r[1:])}
    return qx

def factor_anualidad_cohorte(qx, edad_jub, año_jub, i_tecnico=R_BASE):
    """a_x cohorte: suma v^t * tpx para t=1..(EDAD_MAX_TABLA-edad_jub),
    recorriendo la tabla en diagonal (edad+1, año+1 cada paso)."""
    v = 1.0 / (1.0 + i_tecnico)
    edad, año = edad_jub, año_jub
    tpx, a, t = 1.0, 0.0, 0
    año_max_tabla = max(qx[edad_jub].keys())
    while edad < EDAD_MAX_TABLA:
        q = qx[edad].get(año, qx[edad][año_max_tabla])  # clamp si excede la proyeccion (hasta 2073)
        tpx *= (1.0 - q)
        t += 1
        a += (v ** t) * tpx
        edad += 1
        año += 1
    return a

QX_H = _cargar_qx("36774_H.xlsx", "36774_H")
QX_M = _cargar_qx("36774_M.xlsx", "36774_M")
# Factor unisex: media de qx ponderada por poblacion (exposicion) H/M del
# corte transversal correspondiente -- corregido el 2026-06-26 tras dictamen
# de revisor (punto 2.3): la version anterior usaba una media simple 50/50,
# que el revisor señalo como arbitraria. No se dispone de la composicion real
# de ADOPTANTES del esquema por tramo y sexo (eso sigue siendo limitacion),
# pero si se dispone ya en este punto del script de la poblacion total por
# sexo de cada corte (TOTAL_H/TOTAL_M, cargada mas arriba con el mecanismo
# poblacional homogeneizado), que se usa aqui como ponderador de exposicion:
# una aproximacion mejor que la paridad arbitraria y enteramente trazable a
# datos ya cargados, aunque sigue sin ser la composicion especifica de
# adoptantes (limitacion declarada en Conclusiones).
# Se calcula porque el Tribunal de Justicia de la UE (caso Test-Achats,
# Directiva 2004/113/CE) exige primas y prestaciones unisex en nuevos
# contratos de seguro desde dic-2012; su aplicabilidad exacta a un FPEPE
# español debe discutirse en el texto, no asumirse sin mas.
PESO_H_UNISEX, PESO_M_UNISEX = {}, {}
QX_UNISEX_CORTE = {}
for año_corte in [2023, 2053]:
    w_h = TOTAL_H[año_corte] / (TOTAL_H[año_corte] + TOTAL_M[año_corte])
    w_m = 1.0 - w_h
    PESO_H_UNISEX[año_corte], PESO_M_UNISEX[año_corte] = w_h, w_m
    QX_UNISEX_CORTE[año_corte] = {e: {a: w_h * QX_H[e][a] + w_m * QX_M[e][a] for a in QX_H[e]} for e in QX_H}

FACTOR_RENTA = {}
for año_corte in [2023, 2053]:
    FACTOR_RENTA[año_corte] = {
        "H": factor_anualidad_cohorte(QX_H, EDAD_JUB, año_corte),
        "M": factor_anualidad_cohorte(QX_M, EDAD_JUB, año_corte),
        "unisex": factor_anualidad_cohorte(QX_UNISEX_CORTE[año_corte], EDAD_JUB, año_corte),
    }

print("Factor de conversion a renta vitalicia (cohorte, i_tecnico=2%, jubilacion a los 67):")
for año_corte, d in FACTOR_RENTA.items():
    print(f"  Jubilacion {año_corte}: H={d['H']:.2f}  M={d['M']:.2f}  "
          f"unisex={d['unisex']:.2f} (peso_H={PESO_H_UNISEX[año_corte]:.4f}, peso_M={PESO_M_UNISEX[año_corte]:.4f})"
          f"  (antiguo FA_H=20/FA_M=22 exogeno, no derivado de tablas)")


# ═══════════════════════════════════════════════════════════════════════════
# PIPELINE INTEGRADO: para cada anio, sexo, tramo -> fiscal + fondo
# ═══════════════════════════════════════════════════════════════════════════
def simular_pipeline(pop_dict, delta_t, sexo, r_base=R_BASE):
    rows = []
    fondo = {t: 0.0 for t in range(1, 11)}
    for año in AÑOS:
        for t in range(1, 11):
            N = pop_dict[año][t]
            n_adopt = N * TAKE_UP[t]
            renta = RENTA_MEDIA[t]
            if n_adopt < 1 or renta <= 0:
                rows.append({"sexo": sexo, "año": año, "tramo": t, "N": N,
                             "n_adopt": 0.0, "irpf_perdido": 0.0, "ss_perdido": 0.0,
                             "fiscal_perdido": 0.0, "aporte_fondo": 0.0,
                             "saldo_fondo": fondo[t]})
                continue
            aporte_trab = min(APORT_TRAB * renta, CAP_APORT)
            irpf_u = aporte_trab * TIPO_EFECTIVO_IRPF[t]
            ss_u   = aporte_trab * TASA_SS_TOTAL
            irpf_perdido = irpf_u * n_adopt
            ss_perdido   = ss_u * n_adopt
            aporte_total = aporte_trab * (1 + TAU_EMP) * n_adopt

            fondo[t] = (fondo[t] + aporte_total * delta_t[t]) * (1 + r_base) * (1 - C_GES)

            rows.append({"sexo": sexo, "año": año, "tramo": t, "N": N,
                         "n_adopt": n_adopt, "irpf_perdido": irpf_perdido,
                         "ss_perdido": ss_perdido,
                         "fiscal_perdido": irpf_perdido + ss_perdido,
                         "aporte_fondo": aporte_total, "saldo_fondo": fondo[t]})
    return pd.DataFrame(rows)


print("\nSimulando pipeline Hombres...")
df_h = simular_pipeline(POP_H, DELTA_T_H, "H")
print("Simulando pipeline Mujeres...")
df_m = simular_pipeline(POP_M, DELTA_T_M, "M")
df = pd.concat([df_h, df_m], ignore_index=True)

resumen_anual = df.groupby(["sexo", "año"]).agg(
    N_adopt=("n_adopt", "sum"),
    irpf_perdido=("irpf_perdido", "sum"),
    ss_perdido=("ss_perdido", "sum"),
    fiscal_perdido=("fiscal_perdido", "sum"),
    aporte_fondo=("aporte_fondo", "sum"),
    saldo_fondo=("saldo_fondo", "sum"),
).reset_index()
resumen_anual["coste_fiscal_acumulado"] = resumen_anual.groupby("sexo")["fiscal_perdido"].cumsum() * -1

deficit_total_30a = -resumen_anual["fiscal_perdido"].sum()
fondo_final_2053 = resumen_anual[resumen_anual["año"] == 2053]["saldo_fondo"].sum()
irpf_total = resumen_anual["irpf_perdido"].sum()
ss_total = resumen_anual["ss_perdido"].sum()

print(f"\nCoste fiscal bruto acumulado 30a (H+M): {deficit_total_30a/1e6:,.0f} M EUR")
print(f"  IRPF perdido: {irpf_total/1e6:,.0f} M EUR")
print(f"  SS perdido:   {ss_total/1e6:,.0f} M EUR")
print(f"Fondo privado acumulado a 2053 (H+M): {fondo_final_2053/1e6:,.0f} M EUR")


# ═══════════════════════════════════════════════════════════════════════════
# ANALISIS DISTRIBUTIVO (Gini): microsimulacion individual por corte
# transversal (2023 y 2053), siguiendo la mecanica validada de
# microsim_gini_bav.py. LIMITACION declarada: no es un modelo de carrera
# individual completa (cada corte transversal es independiente), es una
# comparacion de fotos fijas en dos puntos del horizonte temporal.
# ═══════════════════════════════════════════════════════════════════════════
def beta_params(tramo):
    if tramo == 1:
        return None
    _, U = TRAMOS_LIM[tramo]
    if U <= 30_000:
        return 2.0, 5.0
    elif U <= 150_000:
        return 2.5, 4.0
    else:
        return 1.5, 2.0

def generar_rentas(n, tramo, seed):
    if tramo == 1:
        return np.zeros(n, dtype=np.float32)
    L, U = TRAMOS_LIM[tramo]
    if not np.isfinite(U):
        U = TRAMO10_CAP_PRACTICO
    a, b = beta_params(tramo)
    rng = np.random.default_rng(seed)
    x = rng.beta(a, b, size=n).astype(np.float32)
    return (L + (U - L) * x).astype(np.float32)

def calcular_gini(valores):
    v = np.sort(np.asarray(valores, dtype=np.float64))
    n = len(v)
    if v.sum() == 0:
        return 0.0
    idx = np.arange(1, n + 1)
    return (2 * np.sum(idx * v)) / (n * np.sum(v)) - (n + 1) / n

def f_acumulacion(f, n=30):
    if abs(f - 1.0) < 1e-12:
        return float(n)
    return f * (f**n - 1) / (f - 1)

FACTOR = (1 + R_BASE) * (1 - C_GES)
SUMA_30 = f_acumulacion(FACTOR, 30)

def simular_corte_transversal(pop_dict, año, delta_t, FA, sexo, seed_base):
    """Devuelve (pension_sin, pension_con) como arrays float32 concatenados
    de todos los tramos -- sin guardar DataFrame de ~10M filas en memoria,
    para evitar OOM en la VM (3.9GB RAM). Tambien devuelve un resumen por
    tramo (ganancia media) para informe."""
    pensiones_sin, pensiones_con = [], []
    resumen_tramo = []
    for t in range(1, 11):
        n = int(round(pop_dict[año][t]))
        if n <= 0:
            continue
        renta = generar_rentas(n, t, seed=seed_base + t)
        rng_adopt = np.random.default_rng(seed_base + 1000 + t)
        adopta = rng_adopt.random(n) < TAKE_UP[t]

        # LIMITACION declarada (dictamen revisor ronda 2, item 6): esta es una
        # aproximacion simplificada a la pension publica, NO la regla legal
        # completa de la Seguridad Social. Asume carrera completa sin lagunas
        # y tasa de sustitucion del 100% sobre una base reguladora igual a la
        # renta corriente; no aplica topes maximos, pensiones minimas, el
        # computo real sobre los ultimos años cotizados ni las reglas de
        # periodo transitorio. Afecta por igual a "sin" y "con" bAV, por lo
        # que no deberia sesgar de forma relevante las comparaciones relativas
        # (ganancia neta, Gini) entre ambos escenarios.
        pension_sin = (renta / 17.5) * 12.0
        aporte_propio = np.minimum(APORT_TRAB * renta, CAP_APORT)
        base_con = renta - aporte_propio
        pension_pub_con = (base_con / 17.5) * 12.0

        aporte_total = aporte_propio * (1 + TAU_EMP)
        d_t = delta_t[t]
        S_30 = aporte_total * d_t * SUMA_30
        pension_bav_anual = S_30 / FA

        pension_total_con = np.where(adopta, pension_pub_con + pension_bav_anual, pension_sin).astype(np.float32)
        pension_sin = pension_sin.astype(np.float32)

        # --- Estadisticas restringidas a adoptantes (para responder la
        # pregunta: estado vs. pensionista, solo entre quienes
        # adoptan -- el resto no cambia nada por construccion) ---
        n_adopt = int(adopta.sum())
        if n_adopt > 0:
            ahorro_publico_adopt = float((pension_sin[adopta] - pension_pub_con[adopta]).mean())
            ganancia_neta_adopt = float((pension_total_con[adopta] - pension_sin[adopta]).mean())
        else:
            ahorro_publico_adopt = 0.0
            ganancia_neta_adopt = 0.0

        resumen_tramo.append({"sexo": sexo, "tramo": t, "n": n,
                               "pension_sin_media": float(pension_sin.mean()),
                               "pension_con_media": float(pension_total_con.mean()),
                               "ganancia_media": float((pension_total_con - pension_sin).mean()),
                               "pension_sin_p10": float(np.percentile(pension_sin, 10)),
                               "pension_sin_p50": float(np.percentile(pension_sin, 50)),
                               "pension_sin_p90": float(np.percentile(pension_sin, 90)),
                               "pension_con_p10": float(np.percentile(pension_total_con, 10)),
                               "pension_con_p50": float(np.percentile(pension_total_con, 50)),
                               "pension_con_p90": float(np.percentile(pension_total_con, 90)),
                               "n_adopt": n_adopt,
                               "ahorro_publico_anual_medio_por_adoptante": ahorro_publico_adopt,
                               "ganancia_neta_anual_media_por_adoptante": ganancia_neta_adopt})
        pensiones_sin.append(pension_sin)
        pensiones_con.append(pension_total_con)
        del renta, adopta, aporte_propio, base_con, pension_pub_con, aporte_total, S_30, pension_bav_anual

    return (np.concatenate(pensiones_sin), np.concatenate(pensiones_con), pd.DataFrame(resumen_tramo))

print("\nMicrosimulacion individual -- corte transversal 2023...")
sin_h23, con_h23, res_h23 = simular_corte_transversal(POP_H, 2023, DELTA_T_H, FACTOR_RENTA[2023]["H"], "H", seed_base=42)
sin_m23, con_m23, res_m23 = simular_corte_transversal(POP_M, 2023, DELTA_T_M, FACTOR_RENTA[2023]["M"], "M", seed_base=142)
sin_2023 = np.concatenate([sin_h23, sin_m23]); con_2023 = np.concatenate([con_h23, con_m23])
resumen_tramo_2023 = pd.concat([res_h23, res_m23], ignore_index=True)
del sin_h23, con_h23, sin_m23, con_m23; gc.collect()

print("Microsimulacion individual -- corte transversal 2053...")
sin_h53, con_h53, res_h53 = simular_corte_transversal(POP_H, 2053, DELTA_T_H, FACTOR_RENTA[2053]["H"], "H", seed_base=4202)
sin_m53, con_m53, res_m53 = simular_corte_transversal(POP_M, 2053, DELTA_T_M, FACTOR_RENTA[2053]["M"], "M", seed_base=4302)
sin_2053 = np.concatenate([sin_h53, sin_m53]); con_2053 = np.concatenate([con_h53, con_m53])
resumen_tramo_2053 = pd.concat([res_h53, res_m53], ignore_index=True)
del sin_h53, con_h53, sin_m53, con_m53; gc.collect()

gini_2023_sin = calcular_gini(sin_2023)
gini_2023_con = calcular_gini(con_2023)
gini_2053_sin = calcular_gini(sin_2053)
gini_2053_con = calcular_gini(con_2053)
del sin_2023, con_2023, sin_2053, con_2053; gc.collect()

print(f"\nGini pensiones 2023: SIN bAV={gini_2023_sin:.4f}  CON bAV={gini_2023_con:.4f}  (delta {gini_2023_con-gini_2023_sin:+.4f})")
print(f"Gini pensiones 2053: SIN bAV={gini_2053_sin:.4f}  CON bAV={gini_2053_con:.4f}  (delta {gini_2053_con-gini_2053_sin:+.4f})")

print("\nGanancia neta media anual por tramo (2023, EUR/año):")
print(resumen_tramo_2023[["sexo", "tramo", "ganancia_media"]].round(0).to_string(index=False))

# ═══════════════════════════════════════════════════════════════════════════
# ESTADO vs. PENSIONISTA A 2053 (peticion explicita, 2026-06-26):
# tras el mecanismo homogeneizado, ¿esta mejor el sistema a 2053, tanto para
# el Estado (gasto futuro en pension publica) como para el pensionista
# (pension total)? Se calcula solo sobre adoptantes (el resto, por
# construccion, no cambia nada frente al escenario sin bAV).
# ═══════════════════════════════════════════════════════════════════════════
n_adopt_2053_total = float(resumen_tramo_2053["n_adopt"].sum())
ahorro_publico_2053_total = float((resumen_tramo_2053["n_adopt"] * resumen_tramo_2053["ahorro_publico_anual_medio_por_adoptante"]).sum())
ganancia_neta_2053_total = float((resumen_tramo_2053["n_adopt"] * resumen_tramo_2053["ganancia_neta_anual_media_por_adoptante"]).sum())
coste_fiscal_anual_acumulacion = (irpf_total + ss_total) / 30.0

print("\n--- ESTADO vs. PENSIONISTA a 2053 (corte transversal, solo adoptantes) ---")
print(f"Adoptantes en el corte 2053 (H+M): {n_adopt_2053_total:,.0f}")
print(f"Ahorro anual en gasto de pension publica de esos adoptantes (base de cotizacion reducida durante la vida activa): {ahorro_publico_2053_total/1e6:,.1f} M€/año, recurrente mientras vivan jubilados desde 2053")
print(f"Coste fiscal anual medio durante la fase de acumulacion 2023-2053 (IRPF+SS perdido / 30 años, ya reportado): {coste_fiscal_anual_acumulacion/1e6:,.1f} M€/año")
print(f"Ganancia neta total agregada para los adoptantes del corte 2053 (pension total vs. sin bAV): {ganancia_neta_2053_total/1e6:,.1f} M€/año")

# --- Escenario unisex (Test-Achats / Directiva 2004/113/CE): mismo sorteo de
# renta y adopcion (mismas seeds), solo cambia el factor de conversion de
# sexo-especifico a unisex -- aisla el efecto puro del factor actuarial.
print("\nEscenario unisex (mismas seeds, solo cambia el factor de conversion)...")
_, con_h23_u, _ = simular_corte_transversal(POP_H, 2023, DELTA_T_H, FACTOR_RENTA[2023]["unisex"], "H", seed_base=42)
_, con_m23_u, _ = simular_corte_transversal(POP_M, 2023, DELTA_T_M, FACTOR_RENTA[2023]["unisex"], "M", seed_base=142)
con_2023_unisex = np.concatenate([con_h23_u, con_m23_u])
gini_2023_con_unisex = calcular_gini(con_2023_unisex)
print(f"  Gini 2023 CON bAV (factor unisex): {gini_2023_con_unisex:.4f}  "
      f"(vs {gini_2023_con:.4f} con factor por sexo -- diferencia atribuible solo al factor actuarial)")
del con_h23_u, con_m23_u, con_2023_unisex; gc.collect()


# ═══════════════════════════════════════════════════════════════════════════
# SOSTENIBILIDAD SS Y COMPENSACION VIABLE
# ═══════════════════════════════════════════════════════════════════════════
PIB_2023 = 1_498_000  # M EUR (INE/Banco de España, dato publico)
ss_perdido_total = ss_total  # EUR, 30 años acumulados

escenarios = {}
for pct in [0, 25, 50, 75, 100]:
    compensado = ss_perdido_total * (pct / 100)
    balance_final = -(ss_perdido_total - compensado)
    anual_equiv = compensado / 30
    escenarios[pct] = {
        "compensado_M€": round(compensado / 1e6, 0),
        "balance_final_M€": round(balance_final / 1e6, 0),
        "anual_M€": round(anual_equiv / 1e6, 1),
        "pct_PIB_anual": round(anual_equiv / 1e6 / PIB_2023 * 100, 3),
    }

print("\nEscenarios de compensacion SS (sobre SS perdido real calculado, no sobre cifra antigua):")
for pct, d in escenarios.items():
    print(f"  {pct:>3}%: compensado={d['compensado_M€']:>8,.0f} M€  balance={d['balance_final_M€']:>9,.0f} M€  "
          f"anual={d['anual_M€']:>6.1f} M€ ({d['pct_PIB_anual']}% PIB)")


# ═══════════════════════════════════════════════════════════════════════════
# SENSIBILIDAD AL TIPO DE RENTABILIDAD (rentabilidad real neta del fondo,
# determinista, sin estocastica de mercado -- declarado como limitacion: no
# sustituye un analisis de riesgo de mercado/longevidad, solo acota el
# resultado central R_BASE=2% frente a un escenario bajo y uno alto).
# ═══════════════════════════════════════════════════════════════════════════
print("\nSensibilidad del fondo privado 2053 a la rentabilidad real neta (r):")
sensibilidad_fondo = {}
for r_test in [0.010, 0.020, 0.035]:
    df_h_r = simular_pipeline(POP_H, DELTA_T_H, "H", r_base=r_test)
    df_m_r = simular_pipeline(POP_M, DELTA_T_M, "M", r_base=r_test)
    fondo_r = (df_h_r[df_h_r["año"] == 2053]["saldo_fondo"].sum()
               + df_m_r[df_m_r["año"] == 2053]["saldo_fondo"].sum())
    sensibilidad_fondo[r_test] = round(fondo_r / 1e6, 0)
    print(f"  r={r_test:.1%}: fondo 2053 = {fondo_r/1e6:,.0f} M€")
    del df_h_r, df_m_r
gc.collect()


# ═══════════════════════════════════════════════════════════════════════════
# GUARDAR RESULTADOS
# ═══════════════════════════════════════════════════════════════════════════
df.to_csv(OUT / "pipeline_detalle_tramo_año.csv", index=False)
resumen_anual.to_csv(OUT / "resumen_anual.csv", index=False)
resumen_tramo_2023.to_csv(OUT / "resumen_individual_por_tramo_2023.csv", index=False)
resumen_tramo_2053.to_csv(OUT / "resumen_individual_por_tramo_2053.csv", index=False)

resultados = {
    "version": "junio26_v2",
    "fecha": "2026-06-25",
    "horizonte": "2023-2053",
    "supuestos_clave": [
        "TAKE_UP T1-T4 no nulo (2/3/5/7%), resto sin cambios (10/25/40/45/50/50%)",
        "Tipo efectivo IRPF y renta media por tramo: AEAT 2023 real (Tabla505/587), no bracket marginal",
        "Tasa SS total: 28.9% = 28.3% contingencias comunes (tipo legal 2023, no el 31% sin trazar) + 0.6% Mecanismo de Equidad Intergeneracional (MEI, Ley 21/2021, Orden PCM/74/2023), corregido 2026-06-26 (dictamen revisor ronda 3)",
        "Poblacion Hombres y Mujeres: mecanismo UNICO homogeneizado (2026-06-26) -- distribucion real 2023 + salidas reales (Salidas_Anuales) + movilidad P_obs propia de cada sexo, cohorte cerrada, SIN reescalado cruzado entre sexos (sustituye la version anterior, donde Hombres se cargaba tal cual de un Excel con mecanismo distinto -- ver docstring HOMOGENEIZACION)",
        "Gini: dos cortes transversales independientes (2023 y 2053), no carrera individual completa",
        "Factor de conversion a renta vitalicia: calculado por cohorte (no exogeno) a partir de tablas INE 36774_H/M.xlsx, i_tecnico=2%, jubilacion a los 67",
    ],
    "limitaciones_metodologicas": [
        "Identificacion de las matrices de transicion Markov (P_obs): estimadas a partir de marginales agregados por tramo/año, sin datos de panel individual (MCVL u otra fuente longitudinal). No se dispone de la matriz de transicion individual verdadera; P_obs es una aproximacion agregada y se declara como limitacion, no se intenta una nueva identificacion en esta version.",
        "Factor de conversion a renta vitalicia: anualidad inmediata de nivel constante, sin indexacion, sin reversion a beneficiarios, sin recargo de gastos de gestion en la fase de pago (los gastos de gestion ya se aplican en fase de acumulacion via C_GES). El tipo tecnico (2%) se iguala a R_BASE por simplicidad, no es necesariamente el tipo tecnico que aplicaria una aseguradora real.",
        "Rentabilidad del fondo determinista (no estocastica): se acota con un analisis de sensibilidad (ver 'sensibilidad_fondo_2053_M€'), pero no sustituye una simulacion de riesgo de mercado/longevidad con trayectorias aleatorias.",
        "Alcance: simulacion del FPEPE espanol unicamente; no incluye comparacion internacional con otros sistemas (ver decision de alcance 2026-06-25).",
        "Cohorte cerrada: el modelo no incorpora nuevas incorporaciones (jovenes que entran al mercado de trabajo) durante 2023-2053 para ningun sexo; el total poblacional decrece mecanicamente por jubilacion y mortalidad sin reposicion. Limitacion preexistente, ahora simetrica y explicita para Hombres y Mujeres tras la homogeneizacion del 2026-06-26.",
    ],
    "factor_conversion_renta_vitalicia": {
        "metodologia": "Anualidad-cohorte (diagonal edad x año proyectado), i_tecnico=2%, edad jubilacion=67, tabla INE 36774_H/M (qx por mil, 2023-2073)",
        "jubilacion_2023": {k: round(v, 3) for k, v in FACTOR_RENTA[2023].items()},
        "jubilacion_2053": {k: round(v, 3) for k, v in FACTOR_RENTA[2053].items()},
        "gini_2023_con_bAV_factor_unisex": round(float(gini_2023_con_unisex), 4),
        "nota_unisex": "Gini con factor unisex calculado con las mismas seeds que el factor por sexo -- la diferencia frente a gini_2023_con_bAV aisla el efecto puro del factor actuarial (relevante para la discusion legal Test-Achats/Directiva 2004/113/CE)",
    },
    "resultados": {
        "coste_fiscal_bruto_total_M€": round(deficit_total_30a / 1e6, 0),
        "irpf_perdido_total_M€": round(irpf_total / 1e6, 0),
        "ss_perdido_total_M€": round(ss_total / 1e6, 0),
        "fondo_privado_2053_M€": round(fondo_final_2053 / 1e6, 0),
        "pct_PIB_2023_coste_fiscal_acumulado_30a": round(deficit_total_30a / 1e6 / PIB_2023 * 100, 2),
        "pct_PIB_2023_coste_fiscal_anualizado": round((deficit_total_30a / 1e6 / 30) / PIB_2023 * 100, 3),
        "pct_PIB_2023_fondo_acumulado_30a": round(fondo_final_2053 / 1e6 / PIB_2023 * 100, 2),
        "gini_2023_sin_bAV": round(float(gini_2023_sin), 4),
        "gini_2023_con_bAV": round(float(gini_2023_con), 4),
        "gini_2053_sin_bAV": round(float(gini_2053_sin), 4),
        "gini_2053_con_bAV": round(float(gini_2053_con), 4),
    },
    "sensibilidad_fondo_2053_M€": {f"r={k:.1%}": v for k, v in sensibilidad_fondo.items()},
    "escenarios_compensacion_SS": escenarios,
    "analisis_estado_vs_pensionista_2053": {
        "nota": "Solo entre adoptantes del corte transversal 2053; el resto de la poblacion no cambia nada por construccion del modelo.",
        "n_adoptantes_2053": round(n_adopt_2053_total, 0),
        "ahorro_publico_anual_2053_M€": round(ahorro_publico_2053_total / 1e6, 1),
        "coste_fiscal_anual_medio_acumulacion_M€": round(coste_fiscal_anual_acumulacion / 1e6, 1),
        "ganancia_neta_pensionistas_2053_M€_año": round(ganancia_neta_2053_total / 1e6, 1),
    },
}
with open(OUT / "resultados_integrados_junio26.json", "w", encoding="utf-8") as f:
    json.dump(resultados, f, ensure_ascii=False, indent=2)

print(f"\nGuardado en {OUT}")
print("SIMULACION INTEGRADA COMPLETADA")
