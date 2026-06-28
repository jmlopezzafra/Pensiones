#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extension_escenarios.py
========================
Extension de microsim_bav_integrado.py (no se modifica el script original)
para responder a dos puntos pendientes del dictamen de revision (rev02,
PLAN_REVISION_rev02_AIAE_bAV.md, bloques B4):

  B4a. Escenarios de adopcion institucional (no solo voluntario actual):
       - Voluntario actual (TAKE_UP base del modelo, sin cambios)
       - Auto-enrolment con opt-out (toma de referencia: tasas de
         participacion tras inscripcion automatica con posibilidad de
         opt-out, ~85-90% en los esquemas de referencia de la literatura
         de auto-enrolment, Madrian & Shea 2001; Disney et al. 2008) ->
         se aplica 85% de adopcion uniforme en todos los tramos.
       - Obligatorio universal (100% de adopcion en todos los tramos),
         como cota superior teorica.
  B4b. Escenario de estres financiero: rentabilidad real neta negativa
       (-1%), ademas de la horquilla 1%/2%/3,5% ya reportada en el paper.

Reutiliza EXACTAMENTE la misma carga de datos y la misma funcion
simular_pipeline() del script original (mismas fuentes AEAT/INE/TGSS,
mismo R_BASE=2% salvo en el escenario de estres). No se reescriben los
parametros financieros (APORT_TRAB, TAU_EMP, CAP_APORT, C_GES, tipos IRPF,
tasa SS) -- solo se varia TAKE_UP (B4a) y r_base (B4b) sobre el pipeline ya
validado.

ACTUALIZACION 2026-06-26: la carga de poblacion se ha homogeneizado para
usar el mismo mecanismo unico que microsim_bav_integrado.py (cohorte
cerrada, salidas reales restadas, movilidad P_obs propia de cada sexo, sin
reescalado cruzado de Mujeres contra el total de Hombres). Las funciones
de carga se han copiado literalmente del script original para no
modificarlo, tal como exige el parrafo anterior.
"""

import json
from pathlib import Path
import numpy as np
import openpyxl
import pandas as pd

# NOTA DE EMPAQUETADO PARA EL REPOSITORIO PUBLICO (ver la misma nota en
# microsim_bav_integrado.py): rutas relativas a este script segun la
# estructura codigo/ + datos_entrada/ + resultados/ del README.md.
# datos_entrada/01_datos_base/ y 02_poblacion_dinamica/ NO se incluyen
# (fuentes oficiales de terceros, ver README.md); sin ellas este script no
# puede ejecutarse de extremo a extremo, pero el resultado ya calculado
# esta en resultados/extension_escenarios.json.
BASE = Path(__file__).resolve().parent.parent / "datos_entrada"
OUT = Path(__file__).resolve().parent.parent / "resultados"
OUT.mkdir(parents=True, exist_ok=True)

AÑOS = list(range(2023, 2054))

APORT_TRAB = 0.08
TAU_EMP = 0.15
CAP_APORT = 4_704.0
R_BASE = 0.020
C_GES = 0.004
# Correccion 2026-06-26 (dictamen revisor ronda 3): se anade el 0,6% del
# Mecanismo de Equidad Intergeneracional (MEI, Ley 21/2021, Orden PCM/74/2023)
# a la tasa de cotizacion por contingencias comunes (28,3%), igual que en
# microsim_bav_integrado.py, para mantener la coherencia entre ambos scripts.
TASA_SS_TOTAL = 0.283 + 0.006

TAKE_UP_BASE = {1: 0.02, 2: 0.03, 3: 0.05, 4: 0.07,
                5: 0.10, 6: 0.25, 7: 0.40, 8: 0.45, 9: 0.50, 10: 0.50}
TAKE_UP_AUTOENROL = {t: 0.85 for t in range(1, 11)}
TAKE_UP_OBLIGATORIO = {t: 1.00 for t in range(1, 11)}

DELTA_T_H = {1: 0.866, 2: 0.866, 3: 0.866, 4: 0.949, 5: 0.949, 6: 0.949,
             7: 1.000, 8: 1.000, 9: 1.000, 10: 1.000}
DELTA_T_M = {1: 0.837, 2: 0.837, 3: 0.837, 4: 0.938, 5: 0.938, 6: 0.938,
             7: 1.000, 8: 1.000, 9: 1.000, 10: 1.000}

PIB_2023 = 1_498_000  # M EUR


def cargar_irpf_aeat():
    wb505 = openpyxl.load_workbook(BASE / "01_datos_base/Tabla505_2023.xlsx", data_only=True)
    wb587 = openpyxl.load_workbook(BASE / "01_datos_base/Tabla587_2023.xlsx", data_only=True)
    rows505 = list(wb505["Hoja1"].iter_rows(values_only=True))[10:20]
    rows587 = list(wb587["Hoja1"].iter_rows(values_only=True))[10:20]
    renta_media, tipo_efectivo = {}, {}
    for i, t in enumerate(range(1, 11)):
        base = rows505[i][5]
        cuota = rows587[i][5]
        media = rows505[i][7]
        renta_media[t] = max(media, 0.0)
        tipo_efectivo[t] = max(cuota / base, 0.0) if base else 0.0
    return renta_media, tipo_efectivo


RENTA_MEDIA, TIPO_EFECTIVO_IRPF = cargar_irpf_aeat()


# ─── Mecanismo poblacional UNICO y homogeneizado (2026-06-26), IDENTICO al ──
# ─── de microsim_bav_integrado.py: cohorte cerrada, salidas reales restadas, ─
# ─── movilidad P_obs propia de cada sexo, SIN reescalado cruzado entre ─────
# ─── sexos. Se copian aqui las mismas funciones para no modificar el script ─
# ─── original, tal como exige el docstring de este fichero. ────────────────
def _cargar_distribucion_2023_bloques(sh):
    year, v0 = None, {}
    for r in sh.iter_rows(values_only=True):
        if r[0] and isinstance(r[0], str) and r[0].startswith("AÑO"):
            year = int(r[0].split(":")[1].strip())
        elif year == 2023 and isinstance(r[0], int):
            v0[r[0]] = float(r[1])
        elif year is not None and year > 2023:
            break
    return np.array([v0[t] for t in range(1, 11)])


def _cargar_salidas_bloques(sh):
    year, sal = None, {}
    for r in sh.iter_rows(values_only=True):
        if r[0] and isinstance(r[0], str) and r[0].startswith("AÑO"):
            year = int(r[0].split(":")[1].strip())
            sal[year] = {}
        elif isinstance(r[0], int):
            sal[year][r[0]] = float(r[3] or 0.0)
    return sal


def _cargar_distribucion_2023_flat(sh):
    v0 = {}
    for r in sh.iter_rows(min_row=2, values_only=True):
        if r[0] == 2023:
            v0[r[1]] = float(r[3])
    return np.array([v0[t] for t in range(1, 11)])


def _cargar_salidas_flat(sh):
    sal = {}
    for r in sh.iter_rows(min_row=2, values_only=True):
        sal.setdefault(r[0], {})[r[1]] = float(r[5] or 0.0)
    return sal


def _cargar_P_obs(path_excel):
    wbt = openpyxl.load_workbook(path_excel, data_only=True)
    rowsp = list(wbt["P_obs"].iter_rows(min_row=2, max_row=11, max_col=11, values_only=True))
    return np.array([[r[j] for j in range(1, 11)] for r in rowsp])


def propagar_poblacion_markov(v0, sal, P_obs):
    pop = {2023: {t: v0[t - 1] for t in range(1, 11)}}
    v = v0.copy()
    for y in AÑOS[1:]:
        exits = np.array([sal.get(y, {}).get(t, 0.0) for t in range(1, 11)])
        survivors = np.clip(v - exits, 0, None)
        v = survivors @ P_obs
        pop[y] = {t: v[t - 1] for t in range(1, 11)}
    return pop


print("Cargando poblacion (mecanismo homogeneizado, identico a microsim_bav_integrado.py)...")
wb_h = openpyxl.load_workbook(BASE / "02_poblacion_dinamica/Markov_Cotizantes_Hombres_por_Tramo_2023_2053_v1.xlsx", data_only=True)
v0_h = _cargar_distribucion_2023_bloques(wb_h["Cotizantes_por_Tramo"])
sal_h = _cargar_salidas_bloques(wb_h["Salidas_Anuales"])
P_obs_H = _cargar_P_obs(BASE / "03_matrices_transicion/Matrices_Transicion_Hombres.xlsx")

wb_m = openpyxl.load_workbook(BASE / "02_poblacion_dinamica/Markov_Cotizantes_Mujeres_por_Tramo_2023_2053_v1.xlsx", data_only=True)
v0_m = _cargar_distribucion_2023_flat(wb_m["Cotizantes_por_Tramo"])
sal_m = _cargar_salidas_flat(wb_m["Salidas_Anuales"])
P_obs_M = _cargar_P_obs(BASE / "03_matrices_transicion/Matrices_Transicion_Mujeres.xlsx")

POP_H = propagar_poblacion_markov(v0_h, sal_h, P_obs_H)
POP_M = propagar_poblacion_markov(v0_m, sal_m, P_obs_M)


def simular_pipeline(pop_dict, delta_t, sexo, take_up, r_base=R_BASE):
    rows = []
    fondo = {t: 0.0 for t in range(1, 11)}
    for año in AÑOS:
        for t in range(1, 11):
            N = pop_dict[año][t]
            n_adopt = N * take_up[t]
            renta = RENTA_MEDIA[t]
            if n_adopt < 1 or renta <= 0:
                rows.append({"sexo": sexo, "año": año, "tramo": t,
                             "fiscal_perdido": 0.0, "saldo_fondo": fondo[t]})
                continue
            aporte_trab = min(APORT_TRAB * renta, CAP_APORT)
            irpf_u = aporte_trab * TIPO_EFECTIVO_IRPF[t]
            ss_u = aporte_trab * TASA_SS_TOTAL
            irpf_perdido = irpf_u * n_adopt
            ss_perdido = ss_u * n_adopt
            aporte_total = aporte_trab * (1 + TAU_EMP) * n_adopt

            fondo[t] = (fondo[t] + aporte_total * delta_t[t]) * (1 + r_base) * (1 - C_GES)

            rows.append({"sexo": sexo, "año": año, "tramo": t,
                         "fiscal_perdido": irpf_perdido + ss_perdido,
                         "saldo_fondo": fondo[t]})
    return pd.DataFrame(rows)


def total_30a(take_up, r_base=R_BASE):
    df_h = simular_pipeline(POP_H, DELTA_T_H, "H", take_up, r_base)
    df_m = simular_pipeline(POP_M, DELTA_T_M, "M", take_up, r_base)
    df = pd.concat([df_h, df_m], ignore_index=True)
    coste_fiscal_30a = df["fiscal_perdido"].sum()
    fondo_2053 = df[df["año"] == 2053]["saldo_fondo"].sum()
    return coste_fiscal_30a, fondo_2053


resultados = {}

print("\n--- B4a. Escenarios de adopcion institucional (r=2% central) ---")
for nombre, tu in [("voluntario_actual", TAKE_UP_BASE),
                    ("autoenrolment_optout_85pct", TAKE_UP_AUTOENROL),
                    ("obligatorio_universal_100pct", TAKE_UP_OBLIGATORIO)]:
    coste, fondo = total_30a(tu)
    resultados[nombre] = {
        "coste_fiscal_30a_M€": round(coste / 1e6, 0),
        "fondo_2053_M€": round(fondo / 1e6, 0),
        "pct_PIB_coste_acumulado": round(coste / 1e6 / PIB_2023 * 100, 2),
        "pct_PIB_fondo_2053": round(fondo / 1e6 / PIB_2023 * 100, 2),
    }
    print(f"  {nombre:30s}: coste 30a={coste/1e6:>10,.0f} M€  "
          f"fondo 2053={fondo/1e6:>10,.0f} M€")

print("\n--- B4b. Escenario de estres financiero (rentabilidad real negativa) ---")
for r_test in [-0.010, 0.000]:
    coste, fondo = total_30a(TAKE_UP_BASE, r_base=r_test)
    clave = f"estres_r={r_test:.1%}"
    resultados[clave] = {
        "coste_fiscal_30a_M€": round(coste / 1e6, 0),
        "fondo_2053_M€": round(fondo / 1e6, 0),
    }
    print(f"  r={r_test:>6.1%} (take-up base): fondo 2053={fondo/1e6:>10,.0f} M€")

with open(OUT / "extension_escenarios.json", "w", encoding="utf-8") as f:
    json.dump(resultados, f, ensure_ascii=False, indent=2)
print(f"\nGuardado en {OUT / 'extension_escenarios.json'}")
